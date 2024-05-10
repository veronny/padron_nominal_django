from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.shortcuts import render, redirect
from django.db import connection

# filtros
from base.models import DimPeriodo, DimDiscapacidadEtapa, MAESTRO_HIS_ESTABLECIMIENTO
from .models import DimDisFisicaCie,TramaBaseDiscapacidadRpt02FisicaNominal
from django.db.models import Case, When, Value, IntegerField, Sum, OuterRef, Subquery
from django.db.models.functions import Substr, Cast, Concat, Replace
from django.db.models import CharField

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from .utils import generar_operacional

from django.db.models.functions import Substr

# Create your views here.
@login_required
def operacional(request):
    return render(request, 'discapacidad/index.html')

################################################
# REPORTE DE SEGUIMIENTO
################################################
#--- PROVINCIAS -------------------------------------------------------------
def get_provincias(request,provincias_id):
    provincias = (
                 MAESTRO_HIS_ESTABLECIMIENTO
                 .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                 .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                 .values('Provincia','ubigueo_filtrado')
                 .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
              }
    
    return render(request, 'discapacidad/provincias.html', context)


def rpt_operacional_fisico(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        cursor.execute("""
            CREATE TABLE #temp_resultados (
                ubigeo_filtrado VARCHAR(4),
                renaes VARCHAR(20),
                dis_1 INT,
                dis_2 INT,
                dis_3 INT,
                dis_4 INT,
                dis_5 INT,
                dis_6 INT,
                dis_7 INT,
                dis_8 INT,
                dis_9 INT,
                dis_10 INT,
                dis_11 INT,
                dis_12 INT,
                dis_13 INT,
                dis_14 INT,
                dis_15 INT,
                dis_16 INT,
                dis_17 INT,
                dis_18 INT,
                dis_19 INT,
                dis_20 INT,
                dis_21 INT,
                dis_22 INT,
                dis_23 INT,
                dis_24 INT,
                dis_25 INT,
                dis_26 INT,
                dis_27 INT,
                dis_28 INT,
                dis_29 INT,
                dis_30 INT,
                dis_31 INT,
                dis_32 INT,
                dis_33 INT,
                dis_34 INT,
                dis_35 INT,
                dis_36 INT,
                dis_37 INT,
                dis_38 INT,
                dis_39 INT,
                dis_40 INT,
                dis_41 INT,
                dis_42 INT,
                dis_43 INT,
                dis_44 INT,
                dis_45 INT,
                dis_46 INT,
                dis_47 INT,
                dis_48 INT,
                dis_49 INT,
                dis_50 INT,
                dis_51 INT,
                dis_52 INT,
                dis_53 INT,
                dis_54 INT,
                dis_55 INT,
                dis_56 INT,
                dis_57 INT,
                dis_58 INT,
                dis_59 INT,
                dis_60 INT,           
            )
        """)

        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
            INSERT INTO #temp_resultados
            SELECT
                SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                renaes,
                SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_1,
                SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_2,
                SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_3,
                SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_4,
                SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_5,
                SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_6,
                SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_7,
                SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_8,
                SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_9,
                SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_10,
                SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_11,
                SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_12,
                SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_13,
                SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_14,
                SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_15,
                SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_16,
                SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_17,
                SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_18,
                SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_19,
                SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_20,
                SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_21,
                SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_22,
                SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_23,
                SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_24,
                SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_25,
                SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_26,
                SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_27,
                SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_28,
                SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_29,
                SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_30,
                SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_31,
                SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_32,
                SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_33,
                SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_34,
                SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_35,
                SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_36,
                SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_37,
                SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_38,
                SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_39,
                SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_40,
                SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_41,
                SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_42,
                SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_43,
                SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_44,
                SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_45,
                SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_46,
                SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_47,
                SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_48,
                SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_49,
                SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_50,
                SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_51,
                SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_52,
                SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_53,
                SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_54,
                SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_55,
                SUM(CASE WHEN Categoria = 12 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_56,
                SUM(CASE WHEN Categoria = 12 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_57,
                SUM(CASE WHEN Categoria = 12 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_58,
                SUM(CASE WHEN Categoria = 12 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_59,
                SUM(CASE WHEN Categoria = 12 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_60
            FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL
            WHERE SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4) = %s
                AND periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
            GROUP BY SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4), renaes
        """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        cursor.execute("""
            SELECT
                CONCAT(#temp_resultados.ubigeo_filtrado, '-',MAESTRO_HIS_ESTABLECIMIENTO.Provincia) AS nombre_provincia_ubigeo_filtrado,
                #temp_resultados.dis_1,
                #temp_resultados.dis_2,
                #temp_resultados.dis_3,
                #temp_resultados.dis_4,
                #temp_resultados.dis_5,
                #temp_resultados.dis_6,
                #temp_resultados.dis_7,
                #temp_resultados.dis_8,
                #temp_resultados.dis_9,
                #temp_resultados.dis_10,
                #temp_resultados.dis_11,
                #temp_resultados.dis_12,
                #temp_resultados.dis_13,
                #temp_resultados.dis_14,
                #temp_resultados.dis_15,
                #temp_resultados.dis_16,
                #temp_resultados.dis_17,
                #temp_resultados.dis_18,
                #temp_resultados.dis_19,
                #temp_resultados.dis_20,
                #temp_resultados.dis_21,
                #temp_resultados.dis_22,
                #temp_resultados.dis_23,
                #temp_resultados.dis_24,
                #temp_resultados.dis_25,
                #temp_resultados.dis_26,
                #temp_resultados.dis_27,
                #temp_resultados.dis_28,
                #temp_resultados.dis_29,
                #temp_resultados.dis_30,
                #temp_resultados.dis_31,
                #temp_resultados.dis_32,
                #temp_resultados.dis_33,
                #temp_resultados.dis_34,
                #temp_resultados.dis_35,
                #temp_resultados.dis_36,
                #temp_resultados.dis_37,
                #temp_resultados.dis_38,
                #temp_resultados.dis_39,
                #temp_resultados.dis_40,
                #temp_resultados.dis_41,
                #temp_resultados.dis_42,
                #temp_resultados.dis_43,
                #temp_resultados.dis_44,
                #temp_resultados.dis_45,
                #temp_resultados.dis_46,
                #temp_resultados.dis_47,
                #temp_resultados.dis_48,
                #temp_resultados.dis_49,
                #temp_resultados.dis_50,
                #temp_resultados.dis_51,
                #temp_resultados.dis_52,
                #temp_resultados.dis_53,
                #temp_resultados.dis_54,
                #temp_resultados.dis_55,
                #temp_resultados.dis_56,
                #temp_resultados.dis_57,
                #temp_resultados.dis_58,
                #temp_resultados.dis_59,
                #temp_resultados.dis_60
            FROM #temp_resultados
            JOIN MAESTRO_HIS_ESTABLECIMIENTO ON MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico = #temp_resultados.renaes
        """)

        resultado_prov = cursor.fetchall()

    return resultado_prov

def rpt_operacional_sensorial(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        cursor.execute("""
            CREATE TABLE #temp_resultados_sensoriales (
                ubigeo_filtrado VARCHAR(4),
                renaes VARCHAR(20),
                dis_61 INT,
                dis_62 INT,
                dis_63 INT,
                dis_64 INT,
                dis_65 INT,
                dis_66 INT,
                dis_67 INT,
                dis_68 INT,
                dis_69 INT,
                dis_70 INT,
                dis_71 INT,
                dis_72 INT,
                dis_73 INT,
                dis_74 INT,
                dis_75 INT,
                dis_76 INT,
                dis_77 INT,
                dis_78 INT,
                dis_79 INT,
                dis_80 INT,
                dis_81 INT,
                dis_82 INT,
                dis_83 INT,
                dis_84 INT,
                dis_85 INT,
                dis_86 INT,
                dis_87 INT,
                dis_88 INT,
                dis_89 INT,
                dis_90 INT,
                dis_91 INT,
                dis_92 INT,
                dis_93 INT,
                dis_94 INT,
                dis_95 INT,
                dis_96 INT,
                dis_97 INT,
                dis_98 INT,
                dis_99 INT,
                dis_100 INT,
                dis_101 INT,
                dis_102 INT,
                dis_103 INT,
                dis_104 INT,
                dis_105 INT           
            )
        """)

        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
            INSERT INTO #temp_resultados_sensoriales
            SELECT
                SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                renaes,
                SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_61,
                SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_62,
                SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_63,
                SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_64,
                SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_65,
                SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_66,
                SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_67,
                SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_68,
                SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_69,
                SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_70,
                SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_71,
                SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_72,
                SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_73,
                SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_74,
                SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_75,
                SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_76,
                SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_77,
                SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_78,
                SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_79,
                SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_80,
                SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_81,
                SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_82,
                SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_83,
                SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_84,
                SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_85,
                SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_86,
                SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_87,
                SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_88,
                SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_89,
                SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_90,
                SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_91,
                SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_92,
                SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_93,
                SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_94,
                SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_95,
                SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_96,
                SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_97,
                SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_98,
                SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_99,
                SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_100,
                SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_101,
                SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_102,
                SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_103,
                SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_104,
                SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_105
            FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
            WHERE SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4) = %s
                AND periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
            GROUP BY SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4), renaes
        """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        cursor.execute("""
            SELECT
                CONCAT(#temp_resultados_sensoriales.ubigeo_filtrado, '-',MAESTRO_HIS_ESTABLECIMIENTO.Provincia) AS nombre_provincia_ubigeo_filtrado,
                #temp_resultados_sensoriales.dis_61,
                #temp_resultados_sensoriales.dis_62,
                #temp_resultados_sensoriales.dis_63,
                #temp_resultados_sensoriales.dis_64,
                #temp_resultados_sensoriales.dis_65,
                #temp_resultados_sensoriales.dis_66,
                #temp_resultados_sensoriales.dis_67,
                #temp_resultados_sensoriales.dis_68,
                #temp_resultados_sensoriales.dis_69,
                #temp_resultados_sensoriales.dis_70,
                #temp_resultados_sensoriales.dis_71,
                #temp_resultados_sensoriales.dis_72,
                #temp_resultados_sensoriales.dis_73,
                #temp_resultados_sensoriales.dis_74,
                #temp_resultados_sensoriales.dis_75,
                #temp_resultados_sensoriales.dis_76,
                #temp_resultados_sensoriales.dis_77,
                #temp_resultados_sensoriales.dis_78,
                #temp_resultados_sensoriales.dis_79,
                #temp_resultados_sensoriales.dis_80,
                #temp_resultados_sensoriales.dis_81,
                #temp_resultados_sensoriales.dis_82,
                #temp_resultados_sensoriales.dis_83,
                #temp_resultados_sensoriales.dis_84,
                #temp_resultados_sensoriales.dis_85,
                #temp_resultados_sensoriales.dis_86,
                #temp_resultados_sensoriales.dis_87,
                #temp_resultados_sensoriales.dis_88,
                #temp_resultados_sensoriales.dis_89,
                #temp_resultados_sensoriales.dis_90,
                #temp_resultados_sensoriales.dis_91,
                #temp_resultados_sensoriales.dis_92,
                #temp_resultados_sensoriales.dis_93,
                #temp_resultados_sensoriales.dis_94,
                #temp_resultados_sensoriales.dis_95,
                #temp_resultados_sensoriales.dis_96,
                #temp_resultados_sensoriales.dis_97,
                #temp_resultados_sensoriales.dis_98,
                #temp_resultados_sensoriales.dis_99,
                #temp_resultados_sensoriales.dis_100,
                #temp_resultados_sensoriales.dis_101,
                #temp_resultados_sensoriales.dis_102,
                #temp_resultados_sensoriales.dis_103,
                #temp_resultados_sensoriales.dis_104,
                #temp_resultados_sensoriales.dis_105
            FROM #temp_resultados_sensoriales
            JOIN MAESTRO_HIS_ESTABLECIMIENTO ON MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico = #temp_resultados_sensoriales.renaes
        """)

        resultado_prov_sensorial = cursor.fetchall()

    return resultado_prov_sensorial

def rpt_operacional_certificado(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        cursor.execute("""
            CREATE TABLE #temp_resultados_certificado (
                ubigeo_filtrado VARCHAR(4),
                renaes VARCHAR(20),
                dis_106 INT,
                dis_107 INT,
                dis_108 INT,
                dis_109 INT,
                dis_110 INT,
                dis_111 INT,
                dis_112 INT,
                dis_113 INT,
                dis_114 INT,
                dis_115 INT     
            )
        """)

        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
            INSERT INTO #temp_resultados_certificado
            SELECT
                SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                renaes,
                SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_106,
                SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_107,
                SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_108,
                SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_109,
                SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_110,
                SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_111,
                SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_112,
                SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_113,
                SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_114,
                SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_115
            FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL
            WHERE SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4) = %s
                AND periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
            GROUP BY SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4), renaes
        """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        cursor.execute("""
            SELECT
                CONCAT(#temp_resultados_certificado.ubigeo_filtrado, '-',MAESTRO_HIS_ESTABLECIMIENTO.Provincia) AS nombre_provincia_ubigeo_filtrado,
                #temp_resultados_certificado.dis_106,
                #temp_resultados_certificado.dis_107,
                #temp_resultados_certificado.dis_108,
                #temp_resultados_certificado.dis_109,
                #temp_resultados_certificado.dis_110,
                #temp_resultados_certificado.dis_111,
                #temp_resultados_certificado.dis_112,
                #temp_resultados_certificado.dis_113,
                #temp_resultados_certificado.dis_114,
                #temp_resultados_certificado.dis_115
            FROM #temp_resultados_certificado
            JOIN MAESTRO_HIS_ESTABLECIMIENTO ON MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico = #temp_resultados_certificado.renaes
        """)

        resultado_prov_certificado = cursor.fetchall()

    return resultado_prov_certificado

def rpt_operacional_rbc(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        cursor.execute("""
            CREATE TABLE #temp_resultados_rbc (
                ubigeo_filtrado VARCHAR(4),
                renaes VARCHAR(20),
                dis_116 INT,
                dis_117 INT,
                dis_118 INT,
                dis_119 INT,
                dis_120 INT,
                dis_121 INT,
                dis_122 INT,
                dis_123 INT,
                dis_124 INT,
                dis_125 INT,     
                dis_126 INT,    
                dis_127 INT,    
                dis_128 INT,    
                dis_129 INT,    
                dis_130 INT,    
                dis_131 INT,    
                dis_132 INT,
                dis_133 INT,   
                dis_134 INT, 
                dis_135 INT  
            )
        """)

        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
            INSERT INTO #temp_resultados_rbc
            SELECT
                SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                renaes,
                SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_116,
                SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_117,
                SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_118,
                SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_119,
                SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_120,
                SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_121,
                SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_122,
                SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_123,
                SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_124,
                SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_125,
                SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_126,
                SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_127,
                SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_128,
                SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_129,
                SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_130,
                SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_131,
                SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_132,
                SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_133,
                SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_134,
                SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_135
            FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL
            WHERE SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4) = %s
                AND periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
            GROUP BY SUBSTRING(CAST(ubigeo AS VARCHAR(10)), 1, 4), renaes
        """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        cursor.execute("""
            SELECT
                CONCAT(#temp_resultados_rbc.ubigeo_filtrado, '-',MAESTRO_HIS_ESTABLECIMIENTO.Provincia) AS nombre_provincia_ubigeo_filtrado,
                #temp_resultados_rbc.dis_116,
                #temp_resultados_rbc.dis_117,
                #temp_resultados_rbc.dis_118,
                #temp_resultados_rbc.dis_119,
                #temp_resultados_rbc.dis_120,
                #temp_resultados_rbc.dis_121,
                #temp_resultados_rbc.dis_122,
                #temp_resultados_rbc.dis_123,
                #temp_resultados_rbc.dis_124,
                #temp_resultados_rbc.dis_125,
                #temp_resultados_rbc.dis_126,
                #temp_resultados_rbc.dis_127,
                #temp_resultados_rbc.dis_128,
                #temp_resultados_rbc.dis_129,
                #temp_resultados_rbc.dis_130,
                #temp_resultados_rbc.dis_131,
                #temp_resultados_rbc.dis_132,
                #temp_resultados_rbc.dis_133,
                #temp_resultados_rbc.dis_134,
                #temp_resultados_rbc.dis_135        
            FROM #temp_resultados_rbc
            JOIN MAESTRO_HIS_ESTABLECIMIENTO ON MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico = #temp_resultados_rbc.renaes
        """)

        resultado_prov_rbc = cursor.fetchall()

    return resultado_prov_rbc

# validar matriz
def crear_matriz(request):    
    ubigeo = '1201'
    fecha_inicio = '20240102'# Ejemplo de ubigeo
    fecha_fin = '20240110'# Ejemplo de ubigeo
    matriz = rpt_operacional_fisico(ubigeo,fecha_inicio,fecha_fin)  
    # Puedes renderizar la matriz en una plantilla HTML o hacer cualquier otro procesamiento necesario
    return render(request, 'discapacidad/matrizes.html', {'matriz': matriz})

#--- PROVINCIAS EXCEL -------------------------------------------------------------
class RptOperacinalProv(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        provincia = request.GET.get('provincia')

        # Creación de la consulta
        resultado_prov = rpt_operacional_fisico(provincia, fecha_inicio, fecha_fin)
        resultado_prov_sensorial = rpt_operacional_sensorial(provincia, fecha_inicio, fecha_fin)
        resultado_prov_certificado = rpt_operacional_certificado(provincia, fecha_inicio, fecha_fin)
        resultado_prov_rbc = rpt_operacional_rbc(provincia, fecha_inicio, fecha_fin)


        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambia el alto de la columna
        sheet.row_dimensions[1].height = 14
        sheet.row_dimensions[2].height = 14
        sheet.row_dimensions[4].height = 25
        sheet.row_dimensions[15].height = 25
        # cambia el ancho de la columna
        sheet.column_dimensions['A'].width = 2
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 9
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 9
        sheet.column_dimensions['G'].width = 9
        sheet.column_dimensions['H'].width = 9
        sheet.column_dimensions['I'].width = 9
        # linea de division
        sheet.freeze_panes = 'AL8'
        
        # Configuración del fondo y el borde
        fill = PatternFill(patternType='solid', fgColor='00B0F0')
        border = Border(left=Side(style='thin', color='00B0F0'),
                        right=Side(style='thin', color='00B0F0'),
                        top=Side(style='thin', color='00B0F0'),
                        bottom=Side(style='thin', color='00B0F0'))

        borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                right=Side(style='thin', color='A9A9A9'), # Plomo
                top=Side(style='thin', color='A9A9A9'), # Plomo
                bottom=Side(style='thin', color='A9A9A9')) # Plomo

        # crea titulo del reporte
        sheet['B1'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B1'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
        
        sheet['B2'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B2'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
        
        sheet['B4'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B4'].font = Font(name = 'Arial', size= 12, bold = True)
        sheet['B4'] = 'REPORTE DE ACTIVIDADES DEL COMPONENTE DE DISCAPACIDAD'
        
        sheet['B6'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B6'] ='DIRESA / GERESA / DISA:'

        sheet['B7'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B7'] ='HOSPITAL Y/O RED DE SALUD'
        
        sheet['F6'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F6'] ='PERIODO'
        
        sheet['G6'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['G6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['G6'] ='AÑO:'
        
        sheet['G7'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['G7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['G7'] ='MES:'
        
        sheet['H6'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H6'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['H6'].border = borde_plomo
        sheet['H6'] = ''
        
        sheet['H7'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H7'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['H7'].border = borde_plomo
        sheet['H7'] = ''
        
        sheet['B9'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B9'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B9'] ='PERSONAS CON DISCAPACIDAD RECIBEN ATENCION DE REHABILITACION EN ESTABLECIMIENTOS DE SALUD (3000688)'
        
        sheet['B10'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B10'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B10'] ='Capacitación en medicina de rehabilitación integral (5004449)'
        
        sheet['B12'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B12'].font = Font(name = 'Arial', size= 8,)
        sheet['B12'] ='Capacitación' 
        
        sheet['C11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C11'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['C11'].fill = fill
        sheet['C11'].border = border
        sheet['C11'] = 'N°'
        
        sheet['C12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C12'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['C12'].border = borde_plomo
        sheet['C12'] = ''
        
        sheet['D11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D11'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D11'].fill = fill
        sheet['D11'].border = border
        sheet['D11'] = 'Capacitados'
        
        sheet['D12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D12'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D12'].border = borde_plomo
        sheet['D12'] = ''
        
        ########## DISCAPACIDAD FISICA #########################
        
        sheet['B14'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B14'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B14'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Física (5005150)' 
                
        sheet['B15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B15'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B15'].fill = fill
        sheet['B15'].border = border
        sheet['B15'] = 'Atenciones'
        
        sheet['C15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['C15'].fill = fill
        sheet['C15'].border = border
        sheet['C15'] = 'Total'
        
        sheet['D15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D15'].fill = fill
        sheet['D15'].border = border
        sheet['D15'] = 'Niños         (1d - 11a)'
        
        sheet['E15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E15'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['E15'].fill = fill
        sheet['E15'].border = border
        sheet['E15'] = 'Adolescentes (12a - 17a)'
        
        sheet['F15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F15'].fill = fill
        sheet['F15'].border = border
        sheet['F15'] = 'Jóvenes (18a - 29a)'
        
        sheet['G15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G15'].fill = fill
        sheet['G15'].border = border
        sheet['G15'] = 'Adultos (30a - 59a)'
        
        sheet['H15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H15'].fill = fill
        sheet['H15'].border = border
        sheet['H15'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=16, max_row=27, min_col=2, max_col=8):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B16'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B16'].font = Font(name = 'Arial', size= 8)
        sheet['B16'] ='Lesiones medulares (0515001)' 
        
        sheet['B17'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B17'].font = Font(name = 'Arial', size= 8)
        sheet['B17'] ='Amputados de miembro superior (0515002)' 
        
        sheet['B18'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B18'].font = Font(name = 'Arial', size= 8)
        sheet['B18'] ='Amputados de miembro inferior (0515003)' 
        
        sheet['B19'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B19'].font = Font(name = 'Arial', size= 8)
        sheet['B19'] ='Enfermedad muscular y unión mioneural (0515004)' 
        
        sheet['B20'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B20'].font = Font(name = 'Arial', size= 8)
        sheet['B20'] ='Lesiones de nervio periférico (0515005)' 
        
        sheet['B21'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B21'].font = Font(name = 'Arial', size= 8)
        sheet['B21'] ='Trastornos del desarrollo de la función motriz (0515006)' 
        
        sheet['B22'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B22'].font = Font(name = 'Arial', size= 8)
        sheet['B22'] ='Enfermedad articular degenerativa (0515007)' 
        
        sheet['B23'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B23'].font = Font(name = 'Arial', size= 8)
        sheet['B23'] ='Enfermedad cerebro vascular (0515008)' 
        
        sheet['B24'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B24'].font = Font(name = 'Arial', size= 8)
        sheet['B24'] ='Encefalopatía infantil (0515009)' 
        
        sheet['B25'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B25'].font = Font(name = 'Arial', size= 8)
        sheet['B25'] ='Enfermedad de Parkinson (0515010)' 
        
        sheet['B26'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B26'].font = Font(name = 'Arial', size= 8)
        sheet['B26'] ='Síndrome de Down (0515011)' 
        
        sheet['B27'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B27'].font = Font(name = 'Arial', size= 8)
        sheet['B27'] ='Trastornos posturales (0515012)' 
        
        sheet['B28'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B28'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B28'] ='SUB TOTAL' 
        ##########################################################    
        ########## DISCAPACIDAD SENSORIAL ########################
        ##########################################################
        sheet['B30'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B30'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B30'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Sensorial (5005151)' 
                
        sheet['B31'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B31'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B31'].fill = fill
        sheet['B31'].border = border
        sheet['B31'] = 'Atenciones'
        
        sheet['C31'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C31'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['C31'].fill = fill
        sheet['C31'].border = border
        sheet['C31'] = 'Total'
        
        sheet['D31'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D31'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D31'].fill = fill
        sheet['D31'].border = border
        sheet['D31'] = 'Niños         (1d - 11a)'
        
        sheet['E31'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E31'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['E31'].fill = fill
        sheet['E31'].border = border
        sheet['E31'] = 'Adolescentes (12a - 17a)'
        
        sheet['F31'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F31'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F31'].fill = fill
        sheet['F31'].border = border
        sheet['F31'] = 'Jóvenes (18a - 29a)'
        
        sheet['G31'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G31'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G31'].fill = fill
        sheet['G31'].border = border
        sheet['G31'] = 'Adultos (30a - 59a)'
        
        sheet['H31'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H31'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H31'].fill = fill
        sheet['H31'].border = border
        sheet['H31'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=32, max_row=36, min_col=2, max_col=8):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B32'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B32'].font = Font(name = 'Arial', size= 8)
        sheet['B32'] ='Hipoacusia y/o sordera (0515101)' 
        
        sheet['B33'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B33'].font = Font(name = 'Arial', size= 8)
        sheet['B33'] ='Baja visión y/o ceguera (0515102)' 
        
        sheet['B34'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B34'].font = Font(name = 'Arial', size= 8)
        sheet['B34'] ='Sordomudez (0515103)' 
        
        sheet['B35'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B35'].font = Font(name = 'Arial', size= 8)
        sheet['B35'] ='Parálisis cerebral infantil (0515104)' 
        
        sheet['B36'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B36'].font = Font(name = 'Arial', size= 8)
        sheet['B36'] ='Enfermedades cerebro vasculares (0515105)' 
        
        sheet['B37'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B37'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B37'] ='SUB TOTAL' 
        ########################################################
        ########## DISCAPACIDAD MENTAL #########################
        ########################################################
        sheet['B39'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B39'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B39'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Mental (5005152)' 
                
        sheet['B40'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B40'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B40'].fill = fill
        sheet['B40'].border = border
        sheet['B40'] = 'Atenciones'
        
        sheet['C40'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C40'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['C40'].fill = fill
        sheet['C40'].border = border
        sheet['C40'] = 'Total'
        
        sheet['D40'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D40'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D40'].fill = fill
        sheet['D40'].border = border
        sheet['D40'] = 'Niños         (1d - 11a)'
        
        sheet['E40'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E40'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['E40'].fill = fill
        sheet['E40'].border = border
        sheet['E40'] = 'Adolescentes (12a - 17a)'
        
        sheet['F40'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F40'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F40'].fill = fill
        sheet['F40'].border = border
        sheet['F40'] = 'Jóvenes (18a - 29a)'
        
        sheet['G40'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G40'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G40'].fill = fill
        sheet['G40'].border = border
        sheet['G40'] = 'Adultos (30a - 59a)'
        
        sheet['H40'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H40'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H40'].fill = fill
        sheet['H40'].border = border
        sheet['H40'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=41, max_row=44, min_col=2, max_col=8):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B41'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B41'].font = Font(name = 'Arial', size= 8)
        sheet['B41'] ='Trastornos de aprendizaje (0515201)' 
        
        sheet['B42'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B42'].font = Font(name = 'Arial', size= 8)
        sheet['B42'] ='Retraso mental: leve, moderado, severo (0515202)' 
        
        sheet['B43'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B43'].font = Font(name = 'Arial', size= 8)
        sheet['B43'] ='Trastornos del espectro autista (0515203)' 
        
        sheet['B44'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B44'].font = Font(name = 'Arial', size= 8)
        sheet['B44'] ='Otros trastornos de salud mental (0515204)' 
        
        sheet['B45'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B45'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B45'] ='SUB TOTAL' 
        ##################################################
        ########## CERTIFICACION #########################
        ##################################################
        sheet['B47'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B47'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B47'] ='PERSONAS CON DISCAPACIDAD CERTIFICADAS EN ESTABLECIMIENTOS DE SALUD (3000689)' 
                
        sheet['B48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B48'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B48'].fill = fill
        sheet['B48'].border = border
        sheet['B48'] = 'Atenciones'
        
        sheet['C48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C48'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['C48'].fill = fill
        sheet['C48'].border = border
        sheet['C48'] = 'Total'
        
        sheet['D48'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D48'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D48'].fill = fill
        sheet['D48'].border = border
        sheet['D48'] = 'Niños         (1d - 11a)'
        
        sheet['E48'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E48'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['E48'].fill = fill
        sheet['E48'].border = border
        sheet['E48'] = 'Adolescentes (12a - 17a)'
        
        sheet['F48'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F48'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F48'].fill = fill
        sheet['F48'].border = border
        sheet['F48'] = 'Jóvenes (18a - 29a)'
        
        sheet['G48'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G48'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G48'].fill = fill
        sheet['G48'].border = border
        sheet['G48'] = 'Adultos (30a - 59a)'
        
        sheet['H48'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H48'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H48'].fill = fill
        sheet['H48'].border = border
        sheet['H48'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=49, max_row=50, min_col=2, max_col=8):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B49'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B49'].font = Font(name = 'Arial', size= 8)
        sheet['B49'] ='Certificación de Discapacidad (0515204)' 
        
        sheet['B50'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B50'].font = Font(name = 'Arial', size= 8)
        sheet['B50'] ='Certificación de Incapacidad (0515205)' 
        
        sheet['B51'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B51'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B51'] ='SUB TOTAL' 
        ################################################
        ########## VISITAS RBC #########################
        ################################################
        sheet['B53'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B53'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B53'] ='PERSONAS CON DISCAPACIDAD RECIBEN SERVICIOS DE REHABILITACIÓN BASADA EN LA COMUNIDAD (3000690)' 
                
        sheet['B54'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B54'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B54'].fill = fill
        sheet['B54'].border = border
        sheet['B54'] = 'Visitas'
        
        sheet['C54'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C54'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['C54'].fill = fill
        sheet['C54'].border = border
        sheet['C54'] = 'Total'
        
        sheet['D54'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D54'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D54'].fill = fill
        sheet['D54'].border = border
        sheet['D54'] = 'Niños         (1d - 11a)'
        
        sheet['E54'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E54'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['E54'].fill = fill
        sheet['E54'].border = border
        sheet['E54'] = 'Adolescentes (12a - 17a)'
        
        sheet['F54'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F54'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F54'].fill = fill
        sheet['F54'].border = border
        sheet['F54'] = 'Jóvenes (18a - 29a)'
        
        sheet['G54'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G54'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G54'].fill = fill
        sheet['G54'].border = border
        sheet['G54'] = 'Adultos (30a - 59a)'
        
        sheet['H54'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H54'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H54'].fill = fill
        sheet['H54'].border = border
        sheet['H54'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=55, max_row=58, min_col=2, max_col=8):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B55'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B55'].font = Font(name = 'Arial', size= 8)
        sheet['B55'] ='1º Visita' 
        
        sheet['B56'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B56'].font = Font(name = 'Arial', size= 8)
        sheet['B56'] ='2º Visita' 
        
        sheet['B57'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B57'].font = Font(name = 'Arial', size= 8)
        sheet['B57'] ='3º Visita' 
        
        sheet['B58'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B58'].font = Font(name = 'Arial', size= 8)
        sheet['B58'] ='4º a + Visitas' 
        
        sheet['B59'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B59'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B59'] ='SUB TOTAL' 
        
        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B61'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B61'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B61'] ='CAPACITACIÓN A AGENTES COMUNITARIOS EN REHABILITACIÓN BASADA EN LA COMUNIDAD (5005155)' 
                
        sheet['C62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C62'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['C62'].fill = fill
        sheet['C62'].border = border
        sheet['C62'] = '1° TALLER'
        
        sheet['D62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D62'].fill = fill
        sheet['D62'].border = border
        sheet['D62'] = '2° TALLER'
        
        sheet['E62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E62'].fill = fill
        sheet['E62'].border = border
        sheet['E62'] = 'PROMSA'
        
        sheet['C63'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['C63'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['C63'].fill = fill
        sheet['C63'].border = border
        sheet['C63'] = 'N°'
        
        sheet['D63'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D63'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D63'].fill = fill
        sheet['D63'].border = border
        sheet['D63'] = 'Capacitados'
        
        sheet['E63'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E63'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['E63'].fill = fill
        sheet['E63'].border = border
        sheet['E63'] = 'N°'
        
        sheet['F63'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F63'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F63'].fill = fill
        sheet['F63'].border = border
        sheet['F63'] = 'Capacitados'
        
        sheet['G63'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G63'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G63'].fill = fill
        sheet['G63'].border = border
        sheet['G63'] = 'N° '
        
        sheet['H63'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H63'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H63'].fill = fill
        sheet['H63'].border = border
        sheet['H63'] = 'Capacitados'
        
        sheet['G63'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G63'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G63'].fill = fill
        sheet['G63'].border = border
        sheet['G63'] = 'N°'
        
        sheet['H63'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H63'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H63'].fill = fill
        sheet['H63'].border = border
        sheet['H63'] = 'Capacitados'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=55, max_row=58, min_col=2, max_col=8):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
            
        #############################################################################
        #############################################################################                
        # cambina celdas
        sheet.merge_cells('C6:D6')
        sheet.merge_cells('C7:D7')
        
        # Definir ubicaciones específicas para cada columna y su suma total
        columnas_ubicaciones = {
            'PROVINCIA': 'D10',
            'DIS_1': 'D16', 
            'DIS_2': 'E16',
            'DIS_3': 'F16',
            'DIS_4': 'G16',
            'DIS_5': 'H16',
            'DIS_6': 'D17',
            'DIS_7': 'E17',
            'DIS_8': 'F17',
            'DIS_9': 'G17',
            'DIS_10': 'H17',
            'DIS_11': 'D18',
            'DIS_12': 'E18',
            'DIS_13': 'F18',
            'DIS_14': 'G18',
            'DIS_15': 'H18',
            'DIS_16': 'D19',
            'DIS_17': 'E19',
            'DIS_18': 'F19',
            'DIS_19': 'G19',
            'DIS_20': 'H19',
            'DIS_21': 'D20',
            'DIS_22': 'E20',
            'DIS_23': 'F20',
            'DIS_24': 'G20',
            'DIS_25': 'H20',
            'DIS_26': 'D21',
            'DIS_27': 'E21',
            'DIS_28': 'F21',
            'DIS_29': 'G21',
            'DIS_30': 'H21',
            'DIS_31': 'D22',
            'DIS_32': 'E22',
            'DIS_33': 'F22',
            'DIS_34': 'G22',
            'DIS_35': 'H22',
            'DIS_36': 'D23',
            'DIS_37': 'E23',
            'DIS_38': 'F23',
            'DIS_39': 'G23',
            'DIS_40': 'H23',
            'DIS_41': 'D24',
            'DIS_42': 'E24',
            'DIS_43': 'F24',
            'DIS_44': 'G24',
            'DIS_45': 'H24',
            'DIS_46': 'D25',
            'DIS_47': 'E25',
            'DIS_48': 'F25',
            'DIS_49': 'G25',
            'DIS_50': 'H25',
            'DIS_51': 'D26',
            'DIS_52': 'E26',
            'DIS_53': 'F26',
            'DIS_54': 'G26',
            'DIS_55': 'H26',
            'DIS_56': 'D27',
            'DIS_57': 'E27',
            'DIS_58': 'F27',
            'DIS_59': 'G27',
            'DIS_60': 'H27',           
        }
        
        # Definir ubicaciones específicas para cada columna y su suma total
        col_ubi_sensorial = {    
            'PROVINCIA': 'D10',
            'DIS_61': 'D32',
            'DIS_62': 'E32',
            'DIS_63': 'F32',
            'DIS_64': 'G32',
            'DIS_65': 'H32',
            'DIS_66': 'D33',
            'DIS_67': 'E33',
            'DIS_68': 'F33',
            'DIS_69': 'G33',
            'DIS_70': 'H33',
            'DIS_71': 'D34',
            'DIS_72': 'E34',
            'DIS_73': 'F34',
            'DIS_74': 'G34',
            'DIS_75': 'H34',
            'DIS_76': 'D35',
            'DIS_77': 'E35',
            'DIS_78': 'F35',
            'DIS_79': 'G35',
            'DIS_80': 'H35',
            'DIS_81': 'D36',
            'DIS_82': 'E36',
            'DIS_83': 'F36',
            'DIS_84': 'G36',
            'DIS_85': 'H36',
            'DIS_86': 'D41',
            'DIS_87': 'E41',
            'DIS_88': 'F41',
            'DIS_89': 'G41',
            'DIS_90': 'H41',
            'DIS_91': 'D42',
            'DIS_92': 'E42',
            'DIS_93': 'F42',
            'DIS_94': 'G42',
            'DIS_95': 'H42',
            'DIS_96': 'D43',
            'DIS_97': 'E43',
            'DIS_98': 'F43',
            'DIS_99': 'G43',
            'DIS_100': 'H43',
            'DIS_101': 'D44',
            'DIS_102': 'E44',
            'DIS_103': 'F44',
            'DIS_104': 'G44',
            'DIS_105': 'H44'
        }
        
        # Definir ubicaciones específicas para cada columna y su suma total
        col_ubi_certificado = {    
            'PROVINCIA': 'D10',
            'DIS_106': 'D49',
            'DIS_107': 'E49',
            'DIS_108': 'F49',
            'DIS_109': 'G49',
            'DIS_110': 'H49',
            'DIS_111': 'D50',
            'DIS_112': 'E50',
            'DIS_113': 'F50',
            'DIS_114': 'G50',
            'DIS_115': 'H50'
        }
        
        # Definir ubicaciones específicas para cada columna y su suma total
        col_ubi_rbc = {    
            'PROVINCIA': 'D10',
            'DIS_116': 'D55',
            'DIS_117': 'E55',
            'DIS_118': 'F55',
            'DIS_119': 'G55',
            'DIS_120': 'H55',
            'DIS_121': 'D56',
            'DIS_122': 'E56',
            'DIS_123': 'F56',
            'DIS_124': 'G56',
            'DIS_125': 'H56',
            'DIS_126': 'D57',
            'DIS_127': 'E57',
            'DIS_128': 'F57',
            'DIS_129': 'G57',
            'DIS_130': 'H57',
            'DIS_131': 'D58',
            'DIS_132': 'E58',
            'DIS_133': 'F58',
            'DIS_134': 'G58',
            'DIS_135': 'H58'
        }
        
        # Inicializar diccionario para almacenar sumas por columna
        column_sums = {
            'DIS_1': 0,
            'DIS_2': 0,
            'DIS_3': 0,
            'DIS_4': 0,
            'DIS_5': 0,
            'DIS_6': 0,
            'DIS_7': 0,
            'DIS_8': 0,
            'DIS_9': 0,
            'DIS_10': 0,
            'DIS_11': 0,
            'DIS_12': 0,
            'DIS_13': 0,
            'DIS_14': 0,
            'DIS_15': 0,
            'DIS_16': 0,
            'DIS_17': 0,
            'DIS_18': 0,
            'DIS_19': 0,
            'DIS_20': 0,
            'DIS_21': 0,
            'DIS_22': 0,
            'DIS_23': 0,
            'DIS_24': 0,
            'DIS_25': 0,
            'DIS_26': 0,
            'DIS_27': 0,
            'DIS_28': 0,
            'DIS_29': 0,
            'DIS_30': 0,
            'DIS_31': 0,
            'DIS_32': 0,
            'DIS_33': 0,
            'DIS_34': 0,
            'DIS_35': 0,
            'DIS_36': 0,
            'DIS_37': 0,
            'DIS_38': 0,
            'DIS_39': 0,
            'DIS_40': 0,
            'DIS_41': 0,
            'DIS_42': 0,
            'DIS_43': 0,
            'DIS_44': 0,
            'DIS_45': 0,
            'DIS_46': 0,
            'DIS_47': 0,
            'DIS_48': 0,
            'DIS_49': 0,
            'DIS_50': 0,
            'DIS_51': 0,
            'DIS_52': 0,
            'DIS_53': 0,
            'DIS_54': 0,
            'DIS_55': 0,
            'DIS_56': 0,
            'DIS_57': 0,
            'DIS_58': 0,
            'DIS_59': 0,
            'DIS_60': 0,
        }
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_sensorial = {       
            'DIS_61': 0,
            'DIS_62': 0,
            'DIS_63': 0,
            'DIS_64': 0,
            'DIS_65': 0,
            'DIS_66': 0,
            'DIS_67': 0,
            'DIS_68': 0,
            'DIS_69': 0,
            'DIS_70': 0,
            'DIS_71': 0,
            'DIS_72': 0,
            'DIS_73': 0,
            'DIS_74': 0,
            'DIS_75': 0,
            'DIS_76': 0,
            'DIS_77': 0,
            'DIS_78': 0,
            'DIS_79': 0,
            'DIS_80': 0,
            'DIS_81': 0,
            'DIS_82': 0,
            'DIS_83': 0,
            'DIS_84': 0,
            'DIS_85': 0,
            'DIS_86': 0,
            'DIS_87': 0,
            'DIS_88': 0,
            'DIS_89': 0,
            'DIS_90': 0,
            'DIS_91': 0,
            'DIS_92': 0,
            'DIS_93': 0,
            'DIS_94': 0,
            'DIS_95': 0,
            'DIS_96': 0,
            'DIS_97': 0,
            'DIS_98': 0,
            'DIS_99': 0,
            'DIS_100': 0,
            'DIS_101': 0,
            'DIS_102': 0,
            'DIS_103': 0,
            'DIS_104': 0,
            'DIS_105': 0
        }      
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_certificado = {       
            'DIS_106': 0,
            'DIS_107': 0,
            'DIS_108': 0,
            'DIS_109': 0,
            'DIS_110': 0,
            'DIS_111': 0,
            'DIS_112': 0,
            'DIS_113': 0,
            'DIS_114': 0,
            'DIS_115': 0
        }  
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_rbc = {       
            'DIS_116': 0,
            'DIS_117': 0,
            'DIS_118': 0,
            'DIS_119': 0,
            'DIS_120': 0,
            'DIS_121': 0,
            'DIS_122': 0,
            'DIS_123': 0,
            'DIS_124': 0,
            'DIS_125': 0,
            'DIS_126': 0,
            'DIS_127': 0,
            'DIS_128': 0,
            'DIS_129': 0,
            'DIS_130': 0,
            'DIS_131': 0,
            'DIS_132': 0,
            'DIS_133': 0,
            'DIS_134': 0,
            'DIS_135': 0
        }             
        ############################
        ###  DISCAPACIDAD FISICA ###
        ############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov:
            for col_name in column_sums:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(columnas_ubicaciones.keys()).index(col_name)
                    column_sums[col_name] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila: {row}")                        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_name, total_cell in columnas_ubicaciones.items():
            if col_name in column_sums:
                # Obtener la celda correspondiente según la ubicación
                cell = sheet[total_cell]
                # Asignar el valor de la suma a la celda
                cell.value = column_sums[col_name]
                # Aplicar formato a la celda
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        total_sum_cat_1 =  sum([column_sums['DIS_1'], column_sums['DIS_2'], column_sums['DIS_3'],column_sums['DIS_4'],column_sums['DIS_5']])
        total_sum_cat_2 =  sum([column_sums['DIS_6'], column_sums['DIS_7'], column_sums['DIS_8'],column_sums['DIS_9'],column_sums['DIS_10']])
        total_sum_cat_3 =  sum([column_sums['DIS_11'], column_sums['DIS_12'], column_sums['DIS_13'],column_sums['DIS_14'],column_sums['DIS_15']])
        total_sum_cat_4 =  sum([column_sums['DIS_16'], column_sums['DIS_17'], column_sums['DIS_18'],column_sums['DIS_19'],column_sums['DIS_20']])
        total_sum_cat_5 =  sum([column_sums['DIS_21'], column_sums['DIS_22'], column_sums['DIS_23'],column_sums['DIS_24'],column_sums['DIS_25']])
        total_sum_cat_6 =  sum([column_sums['DIS_26'], column_sums['DIS_27'], column_sums['DIS_28'],column_sums['DIS_29'],column_sums['DIS_30']])
        total_sum_cat_7 =  sum([column_sums['DIS_31'], column_sums['DIS_32'], column_sums['DIS_33'],column_sums['DIS_34'],column_sums['DIS_35']])
        total_sum_cat_8 =  sum([column_sums['DIS_36'], column_sums['DIS_37'], column_sums['DIS_38'],column_sums['DIS_39'],column_sums['DIS_40']])
        total_sum_cat_9 =  sum([column_sums['DIS_41'], column_sums['DIS_42'], column_sums['DIS_43'],column_sums['DIS_44'],column_sums['DIS_45']])
        total_sum_cat_10 =  sum([column_sums['DIS_46'], column_sums['DIS_47'], column_sums['DIS_48'],column_sums['DIS_49'],column_sums['DIS_50']])
        total_sum_cat_11 =  sum([column_sums['DIS_51'], column_sums['DIS_52'], column_sums['DIS_53'],column_sums['DIS_54'],column_sums['DIS_55']])
        total_sum_cat_12 =  sum([column_sums['DIS_56'], column_sums['DIS_57'], column_sums['DIS_58'],column_sums['DIS_59'],column_sums['DIS_60']])

        sheet['C16'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C16'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C16'] = total_sum_cat_1     
        
        sheet['C17'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C17'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C17'] = total_sum_cat_2 
        
        sheet['C18'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C18'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C18'] = total_sum_cat_3    
        
        sheet['C19'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C19'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C19'] = total_sum_cat_4    
        
        sheet['C20'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C20'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C20'] = total_sum_cat_5    
        
        sheet['C21'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C21'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C21'] = total_sum_cat_6    
        
        sheet['C22'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C22'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C22'] = total_sum_cat_7    
        
        sheet['C23'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C23'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C23'] = total_sum_cat_8    
        
        sheet['C24'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C24'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C24'] = total_sum_cat_9    
        
        sheet['C25'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C25'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C25'] = total_sum_cat_10 
        
        sheet['C26'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C26'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C26'] = total_sum_cat_11
        
        sheet['C27'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C27'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C27'] = total_sum_cat_12           
       
        # Sumar los valores del VERTICAL      
        total_sum_cat_vertical_1 =  sum([column_sums['DIS_1'],column_sums['DIS_6'],column_sums['DIS_11'],column_sums['DIS_16'],column_sums['DIS_21'],column_sums['DIS_26'],column_sums['DIS_31'],column_sums['DIS_36'],column_sums['DIS_41'],column_sums['DIS_46'],column_sums['DIS_51'],column_sums['DIS_56']])
        total_sum_cat_vertical_2 =  sum([column_sums['DIS_2'],column_sums['DIS_7'],column_sums['DIS_12'],column_sums['DIS_17'],column_sums['DIS_22'],column_sums['DIS_27'],column_sums['DIS_32'],column_sums['DIS_37'],column_sums['DIS_42'],column_sums['DIS_47'],column_sums['DIS_52'],column_sums['DIS_57']])
        total_sum_cat_vertical_3 =  sum([column_sums['DIS_3'],column_sums['DIS_8'], column_sums['DIS_13'],column_sums['DIS_18'],column_sums['DIS_23'],column_sums['DIS_28'],column_sums['DIS_33'],column_sums['DIS_38'],column_sums['DIS_43'],column_sums['DIS_48'],column_sums['DIS_53'],column_sums['DIS_58']])
        total_sum_cat_vertical_4 =  sum([column_sums['DIS_4'],column_sums['DIS_9'],column_sums['DIS_14'],column_sums['DIS_19'],column_sums['DIS_24'],column_sums['DIS_29'],column_sums['DIS_34'],column_sums['DIS_39'],column_sums['DIS_44'],column_sums['DIS_49'],column_sums['DIS_54'],column_sums['DIS_59']])
        total_sum_cat_vertical_5 =  sum([column_sums['DIS_5'],column_sums['DIS_10'],column_sums['DIS_15'],column_sums['DIS_20'],column_sums['DIS_25'],column_sums['DIS_30'],column_sums['DIS_35'],column_sums['DIS_40'],column_sums['DIS_45'],column_sums['DIS_50'],column_sums['DIS_55'],column_sums['DIS_60']])

        sheet['D28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D28'] = total_sum_cat_vertical_1     
        
        sheet['E28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E28'] = total_sum_cat_vertical_2 
        
        sheet['F28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F28'] = total_sum_cat_vertical_3    
        
        sheet['G28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G28'] = total_sum_cat_vertical_4    
        
        sheet['H28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H28'] = total_sum_cat_vertical_5    
        ##########################################################################
        ###############################
        ###  DISCAPACIDAD SENSORIAL ###
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_sensorial:
            for col_sensorial in col_sum_sensorial:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_sensorial.keys()).index(col_sensorial)
                    col_sum_sensorial[col_sensorial] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_sensorial, total_cell_sensorial in col_ubi_sensorial.items():
            if col_sensorial in col_sum_sensorial:
                # Obtener la celda correspondiente según la ubicación
                cell_sensorial = sheet[total_cell_sensorial]
                # Asignar el valor de la suma a la celda
                cell_sensorial.value = col_sum_sensorial[col_sensorial]
                # Aplicar formato a la celda
                cell_sensorial.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_sensorial.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_sensorial.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_sensorial['DIS_61'], col_sum_sensorial['DIS_62'], col_sum_sensorial['DIS_63'], col_sum_sensorial['DIS_64'], col_sum_sensorial['DIS_65']])
        t_sum_cat_2 =  sum([col_sum_sensorial['DIS_66'], col_sum_sensorial['DIS_67'], col_sum_sensorial['DIS_68'], col_sum_sensorial['DIS_69'], col_sum_sensorial['DIS_70']])
        t_sum_cat_3 =  sum([col_sum_sensorial['DIS_71'], col_sum_sensorial['DIS_72'], col_sum_sensorial['DIS_73'], col_sum_sensorial['DIS_74'], col_sum_sensorial['DIS_75']])
        t_sum_cat_4 =  sum([col_sum_sensorial['DIS_76'], col_sum_sensorial['DIS_77'], col_sum_sensorial['DIS_78'], col_sum_sensorial['DIS_79'], col_sum_sensorial['DIS_70']])
        t_sum_cat_5 =  sum([col_sum_sensorial['DIS_81'], col_sum_sensorial['DIS_82'], col_sum_sensorial['DIS_83'], col_sum_sensorial['DIS_84'], col_sum_sensorial['DIS_85']])
        t_sum_cat_6 =  sum([col_sum_sensorial['DIS_86'], col_sum_sensorial['DIS_87'], col_sum_sensorial['DIS_88'], col_sum_sensorial['DIS_89'], col_sum_sensorial['DIS_80']])
        t_sum_cat_7 =  sum([col_sum_sensorial['DIS_91'], col_sum_sensorial['DIS_92'], col_sum_sensorial['DIS_93'], col_sum_sensorial['DIS_94'], col_sum_sensorial['DIS_95']])
        t_sum_cat_8 =  sum([col_sum_sensorial['DIS_96'], col_sum_sensorial['DIS_97'], col_sum_sensorial['DIS_98'], col_sum_sensorial['DIS_99'], col_sum_sensorial['DIS_90']])
        t_sum_cat_9 =  sum([col_sum_sensorial['DIS_101'],col_sum_sensorial['DIS_102'],col_sum_sensorial['DIS_103'],col_sum_sensorial['DIS_104'],col_sum_sensorial['DIS_105']])

        sheet['C32'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C32'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C32'] = t_sum_cat_1     
        
        sheet['C33'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C33'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C33'] = t_sum_cat_2 
        
        sheet['C34'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C34'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C34'] = t_sum_cat_3    
        
        sheet['C35'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C35'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C35'] = t_sum_cat_4    
        
        sheet['C36'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C36'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C36'] = t_sum_cat_5    
        
        sheet['C41'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C41'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C41'] = t_sum_cat_6    
        
        sheet['C42'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C42'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C42'] = t_sum_cat_7    
        
        sheet['C43'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C43'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C43'] = t_sum_cat_8    
        
        sheet['C44'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C44'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C44'] = t_sum_cat_9    
                 
        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_sensorial['DIS_61'],col_sum_sensorial['DIS_66'],col_sum_sensorial['DIS_71'],col_sum_sensorial['DIS_76'],col_sum_sensorial['DIS_81']])
        t_sum_cat_vertical_2 =  sum([col_sum_sensorial['DIS_62'],col_sum_sensorial['DIS_67'],col_sum_sensorial['DIS_72'],col_sum_sensorial['DIS_77'],col_sum_sensorial['DIS_82']])
        t_sum_cat_vertical_3 =  sum([col_sum_sensorial['DIS_63'],col_sum_sensorial['DIS_68'],col_sum_sensorial['DIS_73'],col_sum_sensorial['DIS_78'],col_sum_sensorial['DIS_83']])
        t_sum_cat_vertical_4 =  sum([col_sum_sensorial['DIS_64'],col_sum_sensorial['DIS_69'],col_sum_sensorial['DIS_74'],col_sum_sensorial['DIS_79'],col_sum_sensorial['DIS_84']])
        t_sum_cat_vertical_5 =  sum([col_sum_sensorial['DIS_65'],col_sum_sensorial['DIS_70'],col_sum_sensorial['DIS_75'],col_sum_sensorial['DIS_80'],col_sum_sensorial['DIS_85']])
        
        t_sum_cat_vertical_6 =  sum([col_sum_sensorial['DIS_86'],col_sum_sensorial['DIS_91'],col_sum_sensorial['DIS_96'],col_sum_sensorial['DIS_101']])
        t_sum_cat_vertical_7 =  sum([col_sum_sensorial['DIS_87'],col_sum_sensorial['DIS_92'],col_sum_sensorial['DIS_97'],col_sum_sensorial['DIS_102']])
        t_sum_cat_vertical_8 =  sum([col_sum_sensorial['DIS_88'],col_sum_sensorial['DIS_93'],col_sum_sensorial['DIS_98'],col_sum_sensorial['DIS_103']])
        t_sum_cat_vertical_9 =  sum([col_sum_sensorial['DIS_89'],col_sum_sensorial['DIS_94'],col_sum_sensorial['DIS_99'],col_sum_sensorial['DIS_104']])
        t_sum_cat_vertical_10 = sum([col_sum_sensorial['DIS_90'],col_sum_sensorial['DIS_95'],col_sum_sensorial['DIS_100'],col_sum_sensorial['DIS_105']])

        sheet['D37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D37'] = t_sum_cat_vertical_1     
        
        sheet['E37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E37'] = t_sum_cat_vertical_2 
        
        sheet['F37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F37'] = t_sum_cat_vertical_3    
        
        sheet['G37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G37'] = t_sum_cat_vertical_4    
        
        sheet['H37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H37'] = t_sum_cat_vertical_5    
        
        sheet['D45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D45'] = t_sum_cat_vertical_6     
        
        sheet['E45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E45'] = t_sum_cat_vertical_7 
        
        sheet['F45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F45'] = t_sum_cat_vertical_8    
        
        sheet['G45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G45'] = t_sum_cat_vertical_9    
        
        sheet['H45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H45'] = t_sum_cat_vertical_10    
        ##########################################################################
        
        #################################
        ###  DISCAPACIDAD CERTIFICADO ###
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_certificado:
            for col_certificado in col_sum_certificado:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_certificado.keys()).index(col_certificado)
                    col_sum_certificado[col_certificado] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
             
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_certificado, total_cell_certificado in col_ubi_certificado.items():
            if col_certificado in col_sum_certificado:
                # Obtener la celda correspondiente según la ubicación
                cell_certificado = sheet[total_cell_certificado]
                # Asignar el valor de la suma a la celda
                cell_certificado.value = col_sum_certificado[col_certificado]
                # Aplicar formato a la celda
                cell_certificado.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_certificado.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_certificado.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        # Sumar los valores del diccionario      
        t_sum_cat_cert_1 =  sum([col_sum_certificado['DIS_106'], col_sum_certificado['DIS_107'], col_sum_certificado['DIS_108'], col_sum_certificado['DIS_109'], col_sum_certificado['DIS_110']])
        t_sum_cat_cert_2 =  sum([col_sum_certificado['DIS_111'], col_sum_certificado['DIS_112'], col_sum_certificado['DIS_113'], col_sum_certificado['DIS_114'], col_sum_certificado['DIS_115']])

        sheet['C49'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C49'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C49'] = t_sum_cat_cert_1     
        
        sheet['C50'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C50'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C50'] = t_sum_cat_cert_2 
       
        # Sumar los valores del VERTICAL      
        t_sum_cat_vert_1 =  sum([col_sum_certificado['DIS_106'],col_sum_certificado['DIS_111']])
        t_sum_cat_vert_2 =  sum([col_sum_certificado['DIS_107'],col_sum_certificado['DIS_112']])
        t_sum_cat_vert_3 =  sum([col_sum_certificado['DIS_108'],col_sum_certificado['DIS_113']])
        t_sum_cat_vert_4 =  sum([col_sum_certificado['DIS_109'],col_sum_certificado['DIS_114']])
        t_sum_cat_vert_5 =  sum([col_sum_certificado['DIS_110'],col_sum_certificado['DIS_115']])
        
        sheet['D51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D51'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D51'] = t_sum_cat_vert_1     
        
        sheet['E51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E51'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E51'] = t_sum_cat_vert_2 
        
        sheet['F51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F51'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F51'] = t_sum_cat_vert_3    
        
        sheet['G51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G51'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G51'] = t_sum_cat_vert_4    
        
        sheet['H51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H51'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H51'] = t_sum_cat_vert_5    
        
        #################################
        ###  DISCAPACIDAD RBC ###########
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_rbc:
            for col_rbc in col_sum_rbc:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_rbc.keys()).index(col_rbc)
                    col_sum_rbc[col_rbc] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
             
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_rbc, total_cell_rbc in col_ubi_rbc.items():
            if col_rbc in col_sum_rbc:
                # Obtener la celda correspondiente según la ubicación
                cell_rbc = sheet[total_cell_rbc]
                # Asignar el valor de la suma a la celda
                cell_rbc.value = col_sum_rbc[col_rbc]
                # Aplicar formato a la celda
                cell_rbc.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_rbc.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_rbc.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        
        print(resultado_prov_rbc)
        print(col_rbc)
        print(col_sum_rbc)
        print(col_index)
        
                
        # Sumar los valores del diccionario      
        t_sum_cat_rbc_1 =  sum([col_sum_rbc['DIS_116'], col_sum_rbc['DIS_117'], col_sum_rbc['DIS_118'], col_sum_rbc['DIS_119'], col_sum_rbc['DIS_120']])
        t_sum_cat_rbc_2 =  sum([col_sum_rbc['DIS_121'], col_sum_rbc['DIS_122'], col_sum_rbc['DIS_123'], col_sum_rbc['DIS_124'], col_sum_rbc['DIS_125']])
        t_sum_cat_rbc_3 =  sum([col_sum_rbc['DIS_126'], col_sum_rbc['DIS_127'], col_sum_rbc['DIS_128'], col_sum_rbc['DIS_129'], col_sum_rbc['DIS_130']])
        t_sum_cat_rbc_4 =  sum([col_sum_rbc['DIS_131'], col_sum_rbc['DIS_132'], col_sum_rbc['DIS_133'], col_sum_rbc['DIS_134'], col_sum_rbc['DIS_135']])

        sheet['C55'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C55'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C55'] = t_sum_cat_rbc_1     
        
        sheet['C56'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C56'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C56'] = t_sum_cat_rbc_2 
        
        sheet['C57'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C57'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C57'] = t_sum_cat_rbc_3     
        
        sheet['C58'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C58'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['C58'] = t_sum_cat_rbc_4 
       
        # Sumar los valores del VERTICAL      
        t_sum_vert_rbc_1 =  sum([col_sum_rbc['DIS_116'],col_sum_rbc['DIS_121'],col_sum_rbc['DIS_126'],col_sum_rbc['DIS_131']])
        t_sum_vert_rbc_2 =  sum([col_sum_rbc['DIS_117'],col_sum_rbc['DIS_122'],col_sum_rbc['DIS_127'],col_sum_rbc['DIS_132']])
        t_sum_vert_rbc_3 =  sum([col_sum_rbc['DIS_118'],col_sum_rbc['DIS_123'],col_sum_rbc['DIS_128'],col_sum_rbc['DIS_133']])
        t_sum_vert_rbc_4 =  sum([col_sum_rbc['DIS_119'],col_sum_rbc['DIS_124'],col_sum_rbc['DIS_129'],col_sum_rbc['DIS_134']])
        t_sum_vert_rbc_5 =  sum([col_sum_rbc['DIS_120'],col_sum_rbc['DIS_125'],col_sum_rbc['DIS_130'],col_sum_rbc['DIS_135']])
        
        sheet['D59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D59'] = t_sum_vert_rbc_1     
        
        sheet['E59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E59'] = t_sum_vert_rbc_2 
        
        sheet['F59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F59'] = t_sum_vert_rbc_3    
        
        sheet['G59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G59'] = t_sum_vert_rbc_4    
        
        sheet['H59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H59'] = t_sum_vert_rbc_5  
        
        
         
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_operacional_fisico.xlsx"

        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        workbook.save(response)

        return response

################################################
# REPORTE DE DISTRITO
################################################
def get_distritos(request, distritos_id):
    provincias = (
                 MAESTRO_HIS_ESTABLECIMIENTO
                 .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                 .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                 .values('Provincia','ubigueo_filtrado')
                 .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
              }
     
    return render(request, 'discapacidad/distritos.html', context)

def p_distritos(request):
    provincia_param = request.GET.get('provincia')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct()
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
        
    return render(request, 'discapacidad/partials/p_distritos.html', context)

################################################
# REPORTE POR REDES
################################################
def get_redes(request,redes_id):
    redes = (
                 MAESTRO_HIS_ESTABLECIMIENTO
                 .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                 .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                 .values('Red','codigo_red_filtrado')
                 .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
              }
    
    return render(request, 'discapacidad/redes.html', context)


################################################
# REPORTE POR MICRO-REDES
################################################
def get_microredes(request, microredes_id):
    redes = (
                 MAESTRO_HIS_ESTABLECIMIENTO
                 .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                 .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                 .values('Red','codigo_red_filtrado')
                 .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
              }

    return render(request, 'discapacidad/microredes.html', context)

def p_microredes(request):
    redes_param = request.GET.get('redes')
    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL').distinct()
    # Filtra los establecimientos por el código de la provincia
    if redes_param:
        establecimientos = establecimientos.filter(Codigo_Red__startswith=redes_param[:2])
    # Selecciona el distrito y el código Ubigueo
    microredes = establecimientos.values('MicroRed','Codigo_MicroRed').distinct()  
    context = {
        'redes_param': redes_param,
        'microredes': microredes
    }
        
    return render(request, 'discapacidad/partials/p_microredes.html', context)

#--- ESTABLECIMIENTOS -------------------------------------------------------
def get_establecimientos(request,establecimiento_id):
    redes = (
                 MAESTRO_HIS_ESTABLECIMIENTO
                 .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                 .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                 .values('Red','codigo_red_filtrado')
                 .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
              }
    return render(request,'discapacidad/establecimientos.html', context)


def p_microredes_establec(request):
    redes_param = request.GET.get('redes')
    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Codigo_Red=redes_param)
    # Selecciona el MicroRed y Codigo_MicroRed
    microredes = establecimientos.values('MicroRed','Codigo_MicroRed').distinct()  
    context = {
        'microredes': microredes,
        'is_htmx': True
    }       
    print(establecimientos)
    return render(request, 'discapacidad/partials/p_microredes_establec.html', context)

def p_establecimientos(request):
    microredes = request.GET.get('microredes')
    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')    
    # Filtra los establecimientos por el código de la provincia
    if microredes:
        establecimientos = establecimientos.filter(Codigo_Red__startswith=microredes[:2])
    # Selecciona el distrito y el código Ubigueo
    establec = establecimientos.values('Codigo_Unico','Nombre_Establecimiento').distinct()
    
    context= {
            'establec': establec
             }
    

    return render(request, 'discapacidad/partials/p_establecimientos.html', context)